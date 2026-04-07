#!/usr/bin/env python3
"""
Build year-dependent FLOP link data (2024-2035) from ETYS Appendix B circuit changes.

Approach:
  1. Load existing FLOP 2024 links as curated baseline (from build_flop_data.py)
  2. Load the substation network that produced those links (docs/substation/substation_network.json)
  3. Build per-circuit inventory: which individual circuits belong to each FLOP link?
  4. Load B-2-2 circuit changes, map to FLOP zones via sub_to_flop + propagation
  5. For each year 2024-2035, apply changes and recalculate parallel reactance
  6. Output links_flop_by_year.json

Reactance formula (parallel combination):
  x_eq = 100 / sum(100/x_i)  where x_i is in % on 100 MVA base
  This is the standard per-unit parallel impedance formula.
"""

import sys
import io
import json
import os
from collections import defaultdict
from copy import deepcopy

# Windows encoding fix
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(ROOT)

# ============================================================
# 1. Load existing data
# ============================================================

print("=== Loading data ===")

with open('public/data/links_flop.json') as f:
    flop_links_2024 = json.load(f)
print(f"  FLOP 2024 links: {len(flop_links_2024)}")

with open('scripts/network_to_flop_official.json') as f:
    flop_map = json.load(f)
sub_to_flop = flop_map['sub_to_flop']
print(f"  sub_to_flop mappings: {len(sub_to_flop)}")

with open('docs/substation/substation_network.json') as f:
    network = json.load(f)
branches = network['branches']
print(f"  Substation network branches: {len(branches)}")

# ============================================================
# 2. Verify: reconstruct FLOP 2024 links from substation network
#    This confirms our circuit inventory matches the curated baseline
# ============================================================

print("\n=== Verifying 2024 baseline reconstruction ===")

# Build per-circuit inventory grouped by FLOP link pair
# This mirrors build_flop_data.py lines 158-198
base_circuit_inventory = defaultdict(list)  # link_id -> [{x_pct, rating, sub1, sub2}, ...]

for branch in branches:
    sub1 = branch['sub1']
    sub2 = branch['sub2']
    flop1 = sub_to_flop.get(sub1)
    flop2 = sub_to_flop.get(sub2)

    if flop1 is None or flop2 is None:
        continue
    if flop1 == flop2:
        continue  # Intra-zone circuit

    pair = tuple(sorted([flop1, flop2]))
    link_id = f"{pair[0]}-{pair[1]}"

    x_pct = branch['x_pct']
    rating = branch['rating_mva']

    if x_pct <= 0 or rating <= 0:
        continue

    # Use a counter suffix to make keys unique for double-circuit lines
    base_key = f"{sub1}|{sub2}"
    existing_count = sum(1 for c in base_circuit_inventory[link_id] if c['circuit_key'].startswith(base_key))
    circuit_key = f"{base_key}#{existing_count}" if existing_count > 0 else base_key

    base_circuit_inventory[link_id].append({
        'x_pct': x_pct,
        'rating': rating,
        'sub1': sub1,
        'sub2': sub2,
        'circuit_key': circuit_key
    })

# Verify against existing FLOP 2024 links
existing_by_id = {l['id']: l for l in flop_links_2024}
x_matches = 0
x_mismatches = 0

for link_id, circuits in sorted(base_circuit_inventory.items()):
    x_vals = [c['x_pct'] for c in circuits]
    inv_sum = sum(100.0 / x for x in x_vals if x > 0)
    x_eq = 100.0 / inv_sum if inv_sum > 0 else 999.0
    cap = sum(c['rating'] for c in circuits)

    ex = existing_by_id.get(link_id)
    if ex:
        if abs(x_eq - ex['x_equivalent']) < 0.01:
            x_matches += 1
        else:
            x_mismatches += 1
            print(f"  MISMATCH {link_id}: computed x={x_eq:.4f} vs existing x={ex['x_equivalent']:.4f}")

print(f"  Reactance matches: {x_matches}/{x_matches + x_mismatches}")
if x_mismatches > 0:
    print(f"  WARNING: {x_mismatches} reactance mismatches — check circuit inventory")

# ============================================================
# 3. Load and parse B-2-2 circuit changes
# ============================================================

print("\n=== Loading B-2-2 circuit changes ===")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

wb = openpyxl.load_workbook(
    'docs/neso_docs/ETYS 2024 Appendix-B V1 (1).xlsx',
    read_only=True
)

# ============================================================
# 4. Extend sub_to_flop mapping via circuit-graph propagation
#    New substations in B-2-2 connect to existing ones — inherit FLOP zone
# ============================================================

print("\n=== Extending sub_to_flop mapping ===")

extended_sub_to_flop = dict(sub_to_flop)

# Also propagate from base circuits (B-2-1)
all_circuit_pairs = []
for sheet_name in ['B-2-1a', 'B-2-1b', 'B-2-1c', 'B-2-1d']:
    ws = wb[sheet_name]
    col_offset = 2 if sheet_name == 'B-2-1d' else 0  # OFTO has extra columns
    for row in ws.iter_rows(min_row=3, values_only=True):
        n1_raw = row[0 + col_offset]
        n2_raw = row[1 + col_offset]
        if not n1_raw: continue
        n1 = str(n1_raw).strip()[:4].upper()
        n2 = str(n2_raw).strip()[:4].upper() if n2_raw else ''
        if n1 and n2:
            all_circuit_pairs.append((n1, n2))

for sheet_name in ['B-2-2a', 'B-2-2b', 'B-2-2c']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        n1 = str(row[0]).strip()[:4].upper()
        n2 = str(row[1]).strip()[:4].upper() if row[1] else ''
        if n1 and n2:
            all_circuit_pairs.append((n1, n2))

# Multi-pass propagation
for pass_num in range(5):
    new_count = 0
    for n1, n2 in all_circuit_pairs:
        if n1 in extended_sub_to_flop and n2 not in extended_sub_to_flop:
            extended_sub_to_flop[n2] = extended_sub_to_flop[n1]
            new_count += 1
        elif n2 in extended_sub_to_flop and n1 not in extended_sub_to_flop:
            extended_sub_to_flop[n1] = extended_sub_to_flop[n2]
            new_count += 1
    if new_count == 0:
        break
    print(f"  Pass {pass_num + 1}: propagated {new_count} new mappings")

# Hardcode Western Isles (the 7 remaining unmapped)
western_isles = {'BALA': 'T1', 'HARI': 'T1', 'LEWI': 'T1',
                 'MUAI': 'T1', 'NPIC': 'T1', 'STOR': 'T1', 'TLSK': 'T1'}
for code, zone in western_isles.items():
    if code not in extended_sub_to_flop:
        extended_sub_to_flop[code] = zone

print(f"  Extended mapping: {len(sub_to_flop)} -> {len(extended_sub_to_flop)} substations")

# ============================================================
# 5. Parse B-2-2 changes into structured format
# ============================================================

print("\n=== Parsing B-2-2 circuit changes ===")

circuit_changes = []  # [{year, status, n1, n2, x_pct, rating, flop1, flop2, link_id}, ...]

for sheet_name in ['B-2-2a', 'B-2-2b', 'B-2-2c']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue

        n1_full = str(row[0]).strip()
        n2_full = str(row[1]).strip() if row[1] else ''
        if not n2_full: continue

        year_raw = row[2]
        status = str(row[3]).strip() if row[3] else ''

        # Parse year
        try:
            year = int(year_raw)
        except (ValueError, TypeError):
            continue

        if year < 2025 or year > 2035:
            continue

        # Parse x and rating
        x_raw = row[8]  # X (% on 100MVA)
        rating_raw = row[10]  # Winter Rating (MVA)

        try:
            x_pct = float(x_raw) if x_raw and str(x_raw) != 'TBC' else 0
        except (ValueError, TypeError):
            x_pct = 0

        try:
            rating = float(rating_raw) if rating_raw and str(rating_raw) != 'TBC' else 0
        except (ValueError, TypeError):
            rating = 0

        # Skip bus-couplers
        if rating >= 9999:
            continue

        # Map to FLOP zones
        code1 = n1_full[:4].upper()
        code2 = n2_full[:4].upper()
        flop1 = extended_sub_to_flop.get(code1, '')
        flop2 = extended_sub_to_flop.get(code2, '')

        if not flop1 or not flop2:
            continue  # Can't map — skip
        if flop1 == flop2:
            continue  # Intra-zone — skip

        pair = tuple(sorted([flop1, flop2]))
        link_id = f"{pair[0]}-{pair[1]}"

        circuit_changes.append({
            'year': year,
            'status': status,
            'n1': n1_full,
            'n2': n2_full,
            'code1': code1,
            'code2': code2,
            'x_pct': x_pct,
            'rating': rating,
            'flop1': flop1,
            'flop2': flop2,
            'link_id': link_id,
            'circuit_key': f"{code1}|{code2}|{n1_full}|{n2_full}",
            'sheet': sheet_name
        })

# Also parse B-2-2d (OFTO) — different column layout
ws = wb['B-2-2d']
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[0]: continue
    # OFTO layout: WindFarm, Node1, Node2, Year, Status, ...
    n1_full = str(row[1]).strip() if row[1] else ''
    n2_full = str(row[2]).strip() if row[2] else ''
    if not n1_full or not n2_full: continue

    year_raw = row[3]
    status = str(row[4]).strip() if row[4] else ''

    try:
        year = int(year_raw)
    except (ValueError, TypeError):
        continue

    if year < 2025 or year > 2035:
        continue

    x_raw = row[9]
    rating_raw = row[14] if len(row) > 14 else row[-1]

    try:
        x_pct = float(x_raw) if x_raw and str(x_raw) != 'TBC' else 0
    except (ValueError, TypeError):
        x_pct = 0

    try:
        rating = float(rating_raw) if rating_raw and str(rating_raw) != 'TBC' else 0
    except (ValueError, TypeError):
        rating = 0

    if rating >= 9999:
        continue

    code1 = n1_full[:4].upper()
    code2 = n2_full[:4].upper()
    flop1 = extended_sub_to_flop.get(code1, '')
    flop2 = extended_sub_to_flop.get(code2, '')

    if not flop1 or not flop2 or flop1 == flop2:
        continue

    pair = tuple(sorted([flop1, flop2]))
    link_id = f"{pair[0]}-{pair[1]}"

    circuit_changes.append({
        'year': year,
        'status': status,
        'n1': n1_full,
        'n2': n2_full,
        'code1': code1,
        'code2': code2,
        'x_pct': x_pct,
        'rating': rating,
        'flop1': flop1,
        'flop2': flop2,
        'link_id': link_id,
        'circuit_key': f"{code1}|{code2}|{n1_full}|{n2_full}",
        'sheet': 'B-2-2d'
    })

# Summary
years_count = defaultdict(int)
status_count = defaultdict(int)
for c in circuit_changes:
    years_count[c['year']] += 1
    status_count[c['status']] += 1

print(f"  Total inter-FLOP-zone changes: {len(circuit_changes)}")
print(f"  By year: {dict(sorted(years_count.items()))}")
print(f"  By status: {dict(status_count)}")

affected_links = set(c['link_id'] for c in circuit_changes)
existing_affected = affected_links & set(existing_by_id.keys())
new_links = affected_links - set(existing_by_id.keys())
print(f"  Affected FLOP links: {len(affected_links)} ({len(existing_affected)} existing, {len(new_links)} new)")

# ============================================================
# 6. Build year-by-year FLOP link snapshots
# ============================================================

print("\n=== Building year-by-year FLOP links ===")

# Start with the per-circuit inventory from step 2
# For each year, apply changes cumulatively and recalculate

# Deep copy the base inventory — each circuit is a dict with x_pct, rating, sub1, sub2
# We track circuits by a unique key: sub1|sub2|node1_full|node2_full
# For base circuits, key is sub1|sub2 (from substation_network.json which only has 4-char codes)

# Build mutable circuit state: link_id -> {circuit_key -> {x_pct, rating}}
circuit_state = defaultdict(dict)

for link_id, circuits in base_circuit_inventory.items():
    for c in circuits:
        # Base circuits keyed by sub pair
        key = c['circuit_key']
        circuit_state[link_id][key] = {
            'x_pct': c['x_pct'],
            'rating': c['rating']
        }

def compute_link(circuits_dict):
    """Compute link parameters from a dict of circuits."""
    if not circuits_dict:
        return None
    x_vals = [c['x_pct'] for c in circuits_dict.values() if c['x_pct'] > 0]
    ratings = [c['rating'] for c in circuits_dict.values() if c['rating'] > 0]

    if not x_vals or not ratings:
        return None

    inv_sum = sum(100.0 / x for x in x_vals)
    x_eq = round(100.0 / inv_sum, 4) if inv_sum > 0 else 999.0

    return {
        'x_equivalent': x_eq,
        'capacity_mw': round(sum(ratings), 1),
        'n_circuits': len(x_vals),
    }

# Group changes by year
changes_by_year = defaultdict(list)
for c in circuit_changes:
    changes_by_year[c['year']].append(c)

# Build snapshots
links_by_year = {}

for year in range(2024, 2036):
    # Apply changes for this year (cumulative — changes persist)
    if year in changes_by_year:
        for change in changes_by_year[year]:
            link_id = change['link_id']
            key = change['circuit_key']

            if change['status'] == 'Addition':
                if change['x_pct'] > 0 and change['rating'] > 0:
                    circuit_state[link_id][key] = {
                        'x_pct': change['x_pct'],
                        'rating': change['rating']
                    }
            elif change['status'] == 'Removed':
                # Try exact key match first, then fuzzy by sub pair
                if key in circuit_state[link_id]:
                    del circuit_state[link_id][key]
                else:
                    # Try matching by sub codes (code1|code2)
                    sub_key = f"{change['code1']}|{change['code2']}"
                    candidates = [k for k in circuit_state[link_id] if k.startswith(sub_key) or k == sub_key]
                    if candidates:
                        del circuit_state[link_id][candidates[0]]
            elif change['status'] == 'Change':
                # Update existing circuit's rating (and x if provided)
                if key in circuit_state[link_id]:
                    if change['rating'] > 0:
                        circuit_state[link_id][key]['rating'] = change['rating']
                    if change['x_pct'] > 0:
                        circuit_state[link_id][key]['x_pct'] = change['x_pct']
                else:
                    # Change to a circuit we don't have — treat as addition if valid
                    if change['x_pct'] > 0 and change['rating'] > 0:
                        circuit_state[link_id][key] = {
                            'x_pct': change['x_pct'],
                            'rating': change['rating']
                        }

    # Build link array for this year
    year_links = []
    for link_id in sorted(circuit_state.keys()):
        circuits = circuit_state[link_id]
        if not circuits:
            continue

        result = compute_link(circuits)
        if result is None:
            continue

        parts = link_id.split('-')
        year_links.append({
            'id': link_id,
            'from': parts[0],
            'to': parts[1],
            'x_equivalent': result['x_equivalent'],
            'capacity_mw': result['capacity_mw'],
            'n_circuits': result['n_circuits'],
            'carrier': 'AC'
        })

    links_by_year[str(year)] = year_links

# ============================================================
# 7. Validation
# ============================================================

print("\n=== Validation ===")

# 7a. Check 2024 matches existing
gen_2024 = {l['id']: l for l in links_by_year['2024']}
x_ok = 0
x_bad = 0
for link in flop_links_2024:
    gen = gen_2024.get(link['id'])
    if gen:
        if abs(gen['x_equivalent'] - link['x_equivalent']) < 0.01:
            x_ok += 1
        else:
            x_bad += 1
            print(f"  2024 MISMATCH {link['id']}: gen={gen['x_equivalent']:.4f} vs existing={link['x_equivalent']:.4f}")
    else:
        print(f"  2024 MISSING {link['id']}")

extra = set(gen_2024.keys()) - set(l['id'] for l in flop_links_2024)
if extra:
    print(f"  2024 EXTRA links: {sorted(extra)}")

print(f"  2024 reactance matches: {x_ok}/{x_ok + x_bad}")

# 7b. Year-by-year summary
print(f"\n  Year-by-year summary:")
print(f"  {'Year':<6} {'Links':>6} {'Circuits':>9} {'Capacity':>10} {'x_eq change':>12}")
base_x = {l['id']: l['x_equivalent'] for l in links_by_year['2024']}

for year in range(2024, 2036):
    links = links_by_year[str(year)]
    total_cap = sum(l['capacity_mw'] for l in links)
    total_circuits = sum(l['n_circuits'] for l in links)

    # Count links where reactance changed from 2024
    x_changed = 0
    for l in links:
        if l['id'] in base_x and abs(l['x_equivalent'] - base_x[l['id']]) > 0.001:
            x_changed += 1

    print(f"  {year:<6} {len(links):>6} {total_circuits:>9} {total_cap:>10,.0f} {x_changed:>12}")

# 7c. Check no zero reactances
for year in range(2024, 2036):
    for l in links_by_year[str(year)]:
        if l['x_equivalent'] <= 0:
            print(f"  ERROR: zero reactance {l['id']} in {year}")

# ============================================================
# 8. Output
# ============================================================

output_path = 'public/data/links_flop_by_year.json'
with open(output_path, 'w') as f:
    json.dump(links_by_year, f, indent=2)

file_size = os.path.getsize(output_path)
print(f"\n=== Output ===")
print(f"  Written: {output_path}")
print(f"  Size: {file_size:,} bytes ({file_size // 1024} KB)")
print(f"  Years: {sorted(links_by_year.keys())}")
print("\nDone!")
