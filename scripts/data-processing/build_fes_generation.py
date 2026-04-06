#!/usr/bin/env python3
"""
Extract FES 2025 generation capacity per GSP from BB1, map to zones,
and produce a JSON usable for validation at all resolutions.
"""
import openpyxl, json, sys, io, pandas as pd
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = 'docs/fes/Future Energy Scenarios 2025 Data Workbook V006.xlsx'
PATHWAY = 'Holistic Transition'
YEAR_COL_START = 6  # Column index (0-based) where 2024 starts

# Generation BB mapping: BB ID -> our plant type
GEN_BB_MAP = {
    'Gen_BB014': 'Wind Offshore',
    'Gen_BB015': 'Wind Onshore',
    'Gen_BB016': 'Wind Onshore',  # <1MW, aggregate with onshore
    'Gen_BB012': 'Solar',
    'Gen_BB013': 'Solar',         # Small solar
    'Gen_BB020': 'Nuclear',
    'Gen_BB009': 'CCGT',
    'Gen_BB008': 'OCGT',
    'Gen_BB010': 'Biomass',
    'Gen_BB018': 'Hydro',
    'Gen_BB022': 'Interconnector',
    'Gen_BB001': 'CHP',
    'Gen_BB002': 'CHP',
    'Gen_BB006': 'Gas Engine',
    'Gen_BB005': 'Diesel',
    'Gen_BB023': 'Hydrogen',
    'Gen_BB011': 'Waste',
    'Gen_BB004': 'Renewable Engine',
}

# Demand BB
DEM_BB = {
    'Dem_BB003': 'total_demand_gwh',
    'Dem_BB008': 'baseline_demand_gwh',
}

# Storage BB
STG_BB = {
    'Srg_BB001': 'battery_mw',
    'Srg_BB003': 'pumped_hydro_mw',
}

print(f'Loading {XLSX}...')
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb['BB1']

# Read header row to get year columns
header = None
for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
    header = list(row)
    break

# Find year columns
year_cols = {}
for i, val in enumerate(header):
    if isinstance(val, (int, float)) and 2020 <= val <= 2055:
        year_cols[int(val)] = i

print(f'Year columns found: {sorted(year_cols.keys())[:5]}...{sorted(year_cols.keys())[-3:]}')

# Read all BB1 data for our pathway
gsp_gen = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))  # gsp -> type -> year -> MW
gsp_dem = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))  # gsp -> metric -> year -> value
row_count = 0
gen_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    pathway = row[0]
    bb_id = str(row[1]) if row[1] else ''
    unit = str(row[2]) if row[2] else ''
    gsp = str(row[4]) if row[4] else ''

    if pathway != PATHWAY or not gsp:
        continue

    # Generation (MW only)
    if bb_id in GEN_BB_MAP and unit == 'MW':
        plant_type = GEN_BB_MAP[bb_id]
        for year, col_idx in year_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and isinstance(val, (int, float)):
                gsp_gen[gsp][plant_type][year] += val
                gen_count += 1

    # Demand (GWh)
    if bb_id in DEM_BB and unit == 'GWh':
        metric = DEM_BB[bb_id]
        for year, col_idx in year_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and isinstance(val, (int, float)):
                gsp_dem[gsp][metric][year] += val

    # Storage (MW)
    if bb_id in STG_BB and unit == 'MW':
        metric = STG_BB[bb_id]
        for year, col_idx in year_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val and isinstance(val, (int, float)):
                gsp_gen[gsp][metric][year] += val

    row_count += 1

print(f'Processed {row_count} rows, {gen_count} generation values')
print(f'GSPs with generation: {len(gsp_gen)}')
print(f'GSPs with demand: {len(gsp_dem)}')

# Map GSP to zones (TNUoS, FLOP, Gnode)
sub_zone_map = json.load(open('public/data/substation_zone_mapping.json'))['substations']
flop_map = json.load(open('scripts/network_to_flop_official.json'))['sub_to_flop']
gnode_map = json.load(open('scripts/network_to_gnode.json'))['sub_to_gnode']

# GSP name -> substation code lookup
gsp_to_code = {}
gsp_gnode = pd.read_csv('docs/flop/gsp_gnode_directconnect_region_lookup.csv', encoding='utf-8-sig')
for _, row in gsp_gnode.iterrows():
    gsp_name = str(row.get('gsp_name', ''))
    gnode_name = str(row.get('gnode_name', ''))
    if gsp_name and gsp_name != 'nan':
        # Try matching GSP name first 4 chars to substation code
        code = gsp_name[:4]
        if code in sub_zone_map:
            gsp_to_code[gsp_name] = code

# Also try direct GSP name matching from BB1
for gsp in gsp_gen:
    code = gsp.replace(' ', '')[:4].upper()
    if code in sub_zone_map and gsp not in gsp_to_code:
        gsp_to_code[gsp] = code

# Build FES FLOP lookup from the official CSV
fes_flop = pd.read_csv('docs/flop/fes2022_regional_breakdown_gsp_info.csv', encoding='latin-1')
gsp_to_flop = {}
for _, row in fes_flop.iterrows():
    gsp_id = str(row['GSP ID'])
    flop = row['Minor FLOP']
    if pd.notna(flop):
        gsp_to_flop[gsp_id] = flop
        # Also map the name
        name = str(row['Name'])
        gsp_to_flop[name] = flop

print(f'\nGSP→FLOP mappings from FES CSV: {len(gsp_to_flop)}')

# Aggregate generation by zone for 2024
def aggregate_by_zone(gsp_data, zone_lookup, zone_type='TNUoS'):
    """Aggregate GSP-level data to zone level."""
    zone_gen = defaultdict(lambda: defaultdict(float))
    mapped = 0
    unmapped = []

    for gsp, types in gsp_data.items():
        zone = None

        # Try FES FLOP CSV first (most reliable for FLOP zones)
        if zone_type == 'FLOP':
            zone = gsp_to_flop.get(gsp)
            if not zone:
                # Try GSP ID variants
                for gsp_id, flop in gsp_to_flop.items():
                    if gsp in gsp_id or gsp_id in gsp:
                        zone = flop
                        break

        # Try substation code mapping
        if not zone:
            code = gsp_to_code.get(gsp)
            if code:
                if zone_type == 'TNUoS':
                    zone = sub_zone_map.get(code, {}).get('zone')
                elif zone_type == 'FLOP':
                    zone = flop_map.get(code)
                elif zone_type == 'Gnode':
                    zone = gnode_map.get(code)

        # Try direct name match
        if not zone and zone_type == 'TNUoS':
            for sub_code, sub_data in sub_zone_map.items():
                if gsp.upper() in sub_data.get('name', '').upper():
                    zone = sub_data['zone']
                    break

        if zone:
            for plant_type, years in types.items():
                val = years.get(2024, 0)
                if val > 0:
                    zone_gen[zone][plant_type] += val
            mapped += 1
        else:
            unmapped.append(gsp)

    return dict(zone_gen), mapped, unmapped

# Aggregate for all zone types
for zone_type in ['TNUoS', 'FLOP']:
    zone_gen, mapped, unmapped = aggregate_by_zone(gsp_gen, None, zone_type)
    total_mw = sum(sum(types.values()) for types in zone_gen.values())
    print(f'\n=== FES {PATHWAY} Generation by {zone_type} Zone (2024) ===')
    print(f'Mapped: {mapped}/{len(gsp_gen)} GSPs')
    print(f'Total capacity: {total_mw:.0f} MW across {len(zone_gen)} zones')
    if unmapped[:5]:
        print(f'Unmapped (first 5): {unmapped[:5]}')

    # Show by type
    type_totals = defaultdict(float)
    for zone, types in zone_gen.items():
        for pt, mw in types.items():
            type_totals[pt] += mw
    print(f'By type:')
    for pt, mw in sorted(type_totals.items(), key=lambda x: -x[1]):
        print(f'  {pt}: {mw:.0f} MW')

# Save the FES generation data
output = {
    'metadata': {
        'source': 'FES 2025 Data Workbook BB1',
        'pathway': PATHWAY,
        'year': 2024,
    },
    'generation_by_gsp': {},
    'demand_by_gsp': {},
}

for gsp, types in gsp_gen.items():
    output['generation_by_gsp'][gsp] = {pt: years.get(2024, 0) for pt, years in types.items() if years.get(2024, 0) > 0}

for gsp, metrics in gsp_dem.items():
    output['demand_by_gsp'][gsp] = {m: years.get(2024, 0) for m, years in metrics.items() if years.get(2024, 0) > 0}

with open('scripts/fes_generation_2024.json', 'w') as f:
    json.dump(output, f, indent=2)

print(f'\nSaved to scripts/fes_generation_2024.json')
print(f'GSPs with gen data: {len(output["generation_by_gsp"])}')
print(f'GSPs with demand data: {len(output["demand_by_gsp"])}')

# Compare FES totals vs our TEC Register totals
print(f'\n=== FES vs TEC Register Comparison ===')
plants = json.load(open('public/data/plants_tnuos.json'))
tec_totals = defaultdict(float)
for p in plants:
    if p.get('status') == 'Built' and p.get('mw_connected', 0) > 0:
        pt = p.get('plant_type', '')
        if 'Demand' not in pt and 'Reactive' not in pt and 'Substation' not in pt:
            tec_totals[pt] += p['mw_connected']

print(f'{"Type":<25} {"TEC (MW)":>10} {"FES (MW)":>10} {"Diff":>8}')
all_types = set(list(type_totals.keys()) + list(tec_totals.keys()))
for pt in sorted(all_types):
    tec = tec_totals.get(pt, 0)
    fes = type_totals.get(pt, 0)
    # Try to match FES types to TEC types
    fes_match = 0
    if 'Wind' in pt and 'Offshore' in pt:
        fes_match = type_totals.get('Wind Offshore', 0)
    elif 'Wind' in pt and 'Onshore' in pt:
        fes_match = type_totals.get('Wind Onshore', 0)
    elif 'Solar' in pt or 'PV' in pt:
        fes_match = type_totals.get('Solar', 0)
    elif 'Nuclear' in pt:
        fes_match = type_totals.get('Nuclear', 0)
    elif 'CCGT' in pt:
        fes_match = type_totals.get('CCGT', 0)

    diff = fes - tec if fes > 0 and tec > 0 else 0
    if tec > 0 or fes > 0:
        print(f'{pt:<25} {tec:>10.0f} {fes:>10.0f} {diff:>+8.0f}')
