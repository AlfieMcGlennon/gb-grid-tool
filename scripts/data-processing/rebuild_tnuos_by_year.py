#!/usr/bin/env python3
"""
Rebuild TNUoS year-dependent links with proper reactance recalculation.

Approach: Preserve validated 2024 baseline, apply B-2-2 circuit changes
year-by-year with reactance recalculation via susceptance delta.

For each link at year Y:
  susceptance_Y = susceptance_2024 + sum(1/x_added) - sum(1/x_removed)
  x_equivalent_Y = 1 / susceptance_Y (in % on 100 MVA base)
  capacity_Y = capacity_2024 + sum(rating_added) - sum(rating_removed)

This preserves the curated 2024 reactances while correctly updating
them as circuits change.
"""

import sys
import io
import json
import os
from collections import defaultdict
from copy import deepcopy

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.chdir(ROOT)

# ============================================================
# 1. Load existing data
# ============================================================

print("=== Loading data ===")

with open('public/data/links_tnuos_by_year.json', encoding='utf-8') as f:
    existing = json.load(f)

with open('public/data/substation_zone_mapping.json', encoding='utf-8') as f:
    zone_map = json.load(f)['substations']

existing_2024 = {l['id']: l for l in existing['2024']}
print(f"  Existing 2024 TNUoS links: {len(existing_2024)}")

# ============================================================
# 2. Load and parse B-2-2 circuit changes
# ============================================================

print("\n=== Loading B-2-2 circuit changes ===")

import openpyxl
wb = openpyxl.load_workbook(
    'docs/neso_docs/ETYS 2024 Appendix-B V1 (1).xlsx',
    read_only=True
)

def node_to_tnuos_zone(node_name):
    """Extract 4-char code and look up TNUoS zone."""
    code = str(node_name).strip()[:4].upper()
    return zone_map.get(code, {}).get('zone', ''), code

# Extend zone mapping via circuit-graph propagation (same as FLOP script)
extended_zone_map = {}
for code, info in zone_map.items():
    z = info.get('zone', '')
    if z:
        extended_zone_map[code] = z

all_pairs = []
for sheet_name in ['B-2-1a', 'B-2-1b', 'B-2-1c']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        n1 = str(row[0]).strip()[:4].upper()
        n2 = str(row[1]).strip()[:4].upper() if row[1] else ''
        if n1 and n2:
            all_pairs.append((n1, n2))

for sheet_name in ['B-2-2a', 'B-2-2b', 'B-2-2c']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue
        n1 = str(row[0]).strip()[:4].upper()
        n2 = str(row[1]).strip()[:4].upper() if row[1] else ''
        if n1 and n2:
            all_pairs.append((n1, n2))

for _ in range(5):
    new = 0
    for n1, n2 in all_pairs:
        if n1 in extended_zone_map and n2 not in extended_zone_map:
            extended_zone_map[n2] = extended_zone_map[n1]
            new += 1
        elif n2 in extended_zone_map and n1 not in extended_zone_map:
            extended_zone_map[n1] = extended_zone_map[n2]
            new += 1
    if new == 0:
        break

print(f"  Extended zone mapping: {len(zone_map)} -> {len(extended_zone_map)} substations")

# Parse B-2-2 changes mapped to TNUoS link pairs
circuit_changes = []

for sheet_name in ['B-2-2a', 'B-2-2b', 'B-2-2c']:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]: continue

        n1_full = str(row[0]).strip()
        n2_full = str(row[1]).strip() if row[1] else ''
        if not n2_full: continue

        try:
            year = int(row[2])
        except (ValueError, TypeError):
            continue
        if year < 2025 or year > 2035:
            continue

        status = str(row[3]).strip() if row[3] else ''

        try:
            x_pct = float(row[8]) if row[8] and str(row[8]) != 'TBC' else 0
        except (ValueError, TypeError):
            x_pct = 0

        try:
            rating = float(row[10]) if row[10] and str(row[10]) != 'TBC' else 0
        except (ValueError, TypeError):
            rating = 0

        if rating >= 9999:
            continue  # Bus-coupler

        code1 = n1_full[:4].upper()
        code2 = n2_full[:4].upper()
        z1 = extended_zone_map.get(code1, '')
        z2 = extended_zone_map.get(code2, '')

        if not z1 or not z2 or z1 == z2:
            continue

        pair = tuple(sorted([z1, z2]))
        link_id = f"{pair[0]}-{pair[1]}"

        circuit_changes.append({
            'year': year,
            'status': status,
            'x_pct': x_pct,
            'rating': rating,
            'link_id': link_id,
            'n1': n1_full,
            'n2': n2_full,
        })

# Group by link and year
changes_by_link = defaultdict(lambda: defaultdict(list))
for c in circuit_changes:
    changes_by_link[c['link_id']][c['year']].append(c)

# Filter to only changes affecting existing TNUoS links
relevant_links = set(c['link_id'] for c in circuit_changes if c['link_id'] in existing_2024)
new_links = set(c['link_id'] for c in circuit_changes if c['link_id'] not in existing_2024)

print(f"  Total inter-zone changes: {len(circuit_changes)}")
print(f"  Affecting existing links: {len(relevant_links)}")
print(f"  New link pairs (not in current model): {len(new_links)}")

# ============================================================
# 3. Build year-by-year snapshots with reactance recalculation
# ============================================================

print("\n=== Building year-by-year TNUoS links ===")

links_by_year = {}

for year in range(2024, 2036):
    year_links = []

    for link_id, base_link in sorted(existing_2024.items()):
        # Start from 2024 baseline
        base_susceptance = 100.0 / base_link['x_equivalent']  # Convert x% to susceptance
        base_capacity = base_link['capacity_mw']
        base_circuits = base_link['n_circuits']

        # Accumulate changes from 2025 up to this year
        delta_susceptance = 0
        delta_capacity = 0
        delta_circuits = 0

        for change_year in range(2025, year + 1):
            for change in changes_by_link.get(link_id, {}).get(change_year, []):
                if change['status'] == 'Addition':
                    if change['x_pct'] > 0:
                        delta_susceptance += 100.0 / change['x_pct']
                    if change['rating'] > 0:
                        delta_capacity += change['rating']
                        delta_circuits += 1
                elif change['status'] == 'Removed':
                    if change['x_pct'] > 0:
                        delta_susceptance -= 100.0 / change['x_pct']
                    if change['rating'] > 0:
                        delta_capacity -= change['rating']
                        delta_circuits -= 1
                elif change['status'] == 'Change':
                    # Rating change only — assume x stays same, update capacity
                    if change['rating'] > 0:
                        # Change replaces an existing circuit's rating
                        # We can't perfectly track which circuit changed,
                        # so just update capacity (conservative)
                        delta_capacity += change['rating'] - (base_capacity / max(base_circuits, 1))

        new_susceptance = base_susceptance + delta_susceptance
        new_capacity = base_capacity + delta_capacity
        new_circuits = base_circuits + delta_circuits

        # Guard against zero/negative susceptance
        if new_susceptance <= 0:
            # Link effectively removed (all circuits gone)
            if new_circuits > 0:
                # Circuits remain but susceptance went negative — data issue
                print(f"  WARNING: {link_id} year {year}: susceptance went to {new_susceptance:.2f}, keeping base")
                new_susceptance = base_susceptance
            else:
                continue  # Link genuinely removed

        new_x = 100.0 / new_susceptance

        if new_capacity <= 0 or new_circuits <= 0:
            continue  # Link removed

        year_links.append({
            'id': link_id,
            'from': base_link['from'],
            'to': base_link['to'],
            'capacity_mw': round(new_capacity, 1),
            'n_circuits': new_circuits,
            'x_equivalent': round(new_x, 6),
            'carrier': base_link.get('carrier', 'AC')
        })

    links_by_year[str(year)] = year_links

# ============================================================
# 4. Validation
# ============================================================

print("\n=== Validation ===")

# 4a. 2024 should be identical to existing
gen_2024 = {l['id']: l for l in links_by_year['2024']}
mismatches = 0
for link_id, ex in existing_2024.items():
    gen = gen_2024.get(link_id)
    if not gen:
        print(f"  2024 MISSING: {link_id}")
        mismatches += 1
    elif abs(gen['x_equivalent'] - ex['x_equivalent']) > 0.0001:
        print(f"  2024 x MISMATCH: {link_id} gen={gen['x_equivalent']:.6f} vs ex={ex['x_equivalent']:.6f}")
        mismatches += 1
print(f"  2024 validation: {len(existing_2024) - mismatches}/{len(existing_2024)} match")

# 4b. Year-by-year summary
print(f"\n  {'Year':<6} {'Links':>6} {'Circuits':>9} {'Capacity':>12} {'x changed':>10}")
for year in range(2024, 2036):
    links = links_by_year[str(year)]
    total_cap = sum(l['capacity_mw'] for l in links)
    total_circuits = sum(l['n_circuits'] for l in links)

    x_changed = 0
    for l in links:
        base = existing_2024.get(l['id'])
        if base and abs(l['x_equivalent'] - base['x_equivalent']) > 0.0001:
            x_changed += 1

    print(f"  {year:<6} {len(links):>6} {total_circuits:>9} {total_cap:>12,.0f} {x_changed:>10}")

# 4c. Compare with old frozen-reactance data
print(f"\n  Comparison with old (frozen reactance) data:")
print(f"  {'Year':<6} {'Old cap':>10} {'New cap':>10} {'Old x_chg':>10} {'New x_chg':>10}")
for year in range(2024, 2036):
    old_links = existing.get(str(year), [])
    new_links = links_by_year[str(year)]
    old_cap = sum(l['capacity_mw'] for l in old_links)
    new_cap = sum(l['capacity_mw'] for l in new_links)

    old_x_chg = sum(1 for l in old_links if l['id'] in existing_2024 and abs(l['x_equivalent'] - existing_2024[l['id']]['x_equivalent']) > 0.0001)
    new_x_chg = sum(1 for l in new_links if l['id'] in existing_2024 and abs(l['x_equivalent'] - existing_2024[l['id']]['x_equivalent']) > 0.0001)

    print(f"  {year:<6} {old_cap:>10,.0f} {new_cap:>10,.0f} {old_x_chg:>10} {new_x_chg:>10}")

# 4d. Check no zero reactances
for year in range(2024, 2036):
    for l in links_by_year[str(year)]:
        if l['x_equivalent'] <= 0:
            print(f"  ERROR: zero reactance {l['id']} in {year}")

# ============================================================
# 5. Output
# ============================================================

output_path = 'public/data/links_tnuos_by_year.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(links_by_year, f, indent=2)

file_size = os.path.getsize(output_path)
print(f"\n=== Output ===")
print(f"  Written: {output_path}")
print(f"  Size: {file_size:,} bytes ({file_size // 1024} KB)")
print("\nDone!")
